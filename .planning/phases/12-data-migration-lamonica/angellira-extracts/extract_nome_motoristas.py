#!/usr/bin/env python3
"""Extrai a aba MOTORISTAS de 'nome dos motoristas.xlsx' (elo de identidade
Shopee driver_id <-> CPF <-> nome <-> placa <-> CNH) para JSON normalizado.

Saida: nome_motoristas.json (gitignored - PII).
Chaves normalizadas: driver_id (str), cpf (11 digitos).
"""
import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl ausente: pip install openpyxl")

ROOT = Path(r"C:\Users\antonio.magalhaes\Documents\Projetos\argon-dashboard\.claude\worktrees\elastic-napier-5559df")
XLSX = ROOT / "nome dos motoristas.xlsx"
OUT = Path(__file__).parent / "nome_motoristas.json"

def digits(v):
    if v is None:
        return None
    d = re.sub(r"\D", "", str(v))
    return d or None

def cpf11(v):
    d = digits(v)
    if not d:
        return None
    return d.zfill(11)[-11:] if len(d) <= 11 else d

def norm_id(v):
    """Shopee driver_id pode vir como '2.584.382' ou 2584382.0 -> '2584382'."""
    if v is None:
        return None
    s = re.sub(r"\D", "", str(v))
    return s.lstrip("0") or s or None

def s(v):
    if v is None:
        return None
    t = str(v).strip()
    return t or None

def main():
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    if "MOTORISTAS" not in wb.sheetnames:
        sys.exit(f"aba MOTORISTAS nao encontrada. Abas: {wb.sheetnames}")
    ws = wb["MOTORISTAS"]
    rows = ws.iter_rows(values_only=True)
    header = [s(h) for h in next(rows)]
    # mapeia indices por nome (case-insensitive, contains)
    def idx(*needles):
        for i, h in enumerate(header):
            if h and all(n.lower() in h.lower() for n in needles):
                return i
        return None
    I = {
        "driver_id": idx("driver", "id"),
        "name": idx("driver", "name") or idx("name"),
        "cpf": idx("cpf"),
        "plate": idx("plate"),
        "phone": idx("phone"),
        "vehicle_type": idx("vehicle", "type"),
        "cnh": idx("license", "number"),
        "cnh_cat": idx("license", "type"),
        "cnh_exp": idx("license", "expiry"),
        "station": idx("station"),
        "city": idx("city"),
        "dob": idx("birth"),
        "renavam": idx("renavam"),
        "status": idx("status"),
        "driver_type": idx("driver", "type"),
        "default_rate": idx("default", "rate"),
        "new_rate": idx("new", "rate"),
    }
    out = []
    seen_id, seen_cpf = set(), set()
    for r in rows:
        if r is None or all(c is None for c in r):
            continue
        def g(k):
            i = I[k]
            return r[i] if (i is not None and i < len(r)) else None
        did = norm_id(g("driver_id"))
        cpf = cpf11(g("cpf"))
        if not did and not cpf:
            continue
        rec = {
            "driver_id": did,
            "cpf": cpf,
            "name": s(g("name")),
            "plate": s(g("plate")),
            "phone": s(g("phone")),
            "vehicle_type": s(g("vehicle_type")),
            "cnh": digits(g("cnh")),
            "cnh_cat": s(g("cnh_cat")),
            "cnh_exp": s(g("cnh_exp")),
            "station": s(g("station")),
            "city": s(g("city")),
            "dob": s(g("dob")),
            "renavam": digits(g("renavam")),
            "status": s(g("status")),
            "driver_type": s(g("driver_type")),
            "default_rate": g("default_rate"),
            "new_rate": g("new_rate"),
        }
        out.append(rec)
        if did:
            seen_id.add(did)
        if cpf:
            seen_cpf.add(cpf)
    OUT.write_text(json.dumps(out, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"header_cols={len(header)}")
    print(f"index_map={ {k:v for k,v in I.items()} }")
    print(f"linhas={len(out)} driver_ids_unicos={len(seen_id)} cpfs_unicos={len(seen_cpf)}")
    with_both = sum(1 for r in out if r['driver_id'] and r['cpf'])
    print(f"com_driver_id_e_cpf={with_both}")
    print("amostra=", json.dumps(out[:2], ensure_ascii=False))

if __name__ == "__main__":
    main()
